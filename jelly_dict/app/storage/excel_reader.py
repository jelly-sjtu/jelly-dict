"""Read-side helpers for the vocabulary Excel file.

These functions never mutate the workbook on disk. For mutations see
`app.storage.excel_writer`.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from app.core.models import VocabularyEntry, normalize_word_key
from app.storage.excel_serializer import (
    SHEET_NAME,
    label_to_key,
    row_to_entry,
)


def list_entries(path: Path) -> list[VocabularyEntry]:
    """Return one VocabularyEntry per data row in the workbook."""
    if not path.exists():
        return []
    try:
        wb = load_workbook(path, read_only=True)
    except Exception:
        return []
    if SHEET_NAME not in wb.sheetnames:
        return []
    ws = wb[SHEET_NAME]
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return []
    keys = [label_to_key(label) for label in header]
    out: list[VocabularyEntry] = []
    for raw in rows:
        if raw is None or not any(c is not None for c in raw):
            continue
        out.append(row_to_entry(keys, raw))
    return out


def find_existing(path: Path, language: str, word_key: str) -> VocabularyEntry | None:
    """Return a thin VocabularyEntry from the row matching (language, word_key).

    The Excel sheet is the source of truth for data the user may have
    edited by hand, so we read the row instead of the cache.
    """
    if not path.exists():
        return None
    try:
        wb = load_workbook(path, read_only=True)
    except Exception:
        return None
    if SHEET_NAME not in wb.sheetnames:
        return None
    ws = wb[SHEET_NAME]
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return None
    keys = [label_to_key(label) for label in header]
    if "language" not in keys or "word" not in keys:
        return None
    lang_idx = keys.index("language")
    word_idx = keys.index("word")

    for raw in rows:
        if raw is None or len(raw) <= max(lang_idx, word_idx):
            continue
        row_lang = (raw[lang_idx] or "").strip() if isinstance(raw[lang_idx], str) else ""
        row_word = (raw[word_idx] or "").strip() if isinstance(raw[word_idx], str) else ""
        if row_lang != language:
            continue
        if normalize_word_key(row_word, language) != word_key:  # type: ignore[arg-type]
            continue
        return row_to_entry(keys, raw)
    return None


def read_header(ws) -> list[str]:
    """Return the canonical field-key list inferred from row 1."""
    if ws.max_row < 1:
        return []
    labels = [
        ws.cell(row=1, column=idx).value
        for idx in range(1, ws.max_column + 1)
    ]
    if not any(labels):
        return []
    return [label_to_key(label) for label in labels]

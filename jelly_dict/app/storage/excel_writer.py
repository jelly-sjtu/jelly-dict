"""Write-side helpers for the vocabulary Excel file.

Reader and serializer concerns live in `excel_reader.py` and
`excel_serializer.py` respectively. For backward compatibility this
module re-exports their public symbols so existing callers like
``from app.storage import excel_writer; excel_writer.list_entries(...)``
keep working without import changes.
"""
from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from app.core.errors import ExcelFormatError, ExcelLockedError, StorageError
from app.core.models import VocabularyEntry, normalize_word_key
from app.storage.excel_reader import find_existing, list_entries, read_header
from app.storage.excel_serializer import (
    COLUMN_LABELS,
    COLUMN_WIDTHS,
    HEADER_FILL,
    HEADER_FONT,
    SHEET_NAME,
    label_to_key as _label_to_key,
    render_cell as _render_cell,
    row_to_entry as _row_to_entry,
)

log = logging.getLogger(__name__)

# Re-exports: keep the existing public surface identical.
__all__ = [
    "SHEET_NAME",
    "COLUMN_LABELS",
    "COLUMN_WIDTHS",
    "HEADER_FILL",
    "HEADER_FONT",
    "ensure_workbook",
    "append_entry",
    "update_or_append",
    "delete_entries",
    "save_with_resolver",
    "list_entries",
    "find_existing",
]


def ensure_workbook(path: Path, columns: list[str]) -> None:
    """Create the file with header row if it does not already exist."""
    if path.exists():
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    _write_header(ws, columns)
    _save(wb, path)


def append_entry(path: Path, entry: VocabularyEntry, columns: list[str]) -> None:
    """Append a single entry. Creates the file if missing."""
    if not path.exists():
        ensure_workbook(path, columns)

    wb = _load_for_write(path)
    if SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_NAME)
        _write_header(ws, columns)
    else:
        ws = wb[SHEET_NAME]

    file_columns = read_header(ws) or columns
    if not read_header(ws):
        _write_header(ws, file_columns)

    row_values = [_render_cell(entry, key) for key in file_columns]
    ws.append(row_values)
    _style_last_row(ws, file_columns)
    _save(wb, path)


def update_or_append(path: Path, entry: VocabularyEntry, columns: list[str]) -> None:
    """Replace an existing row with the same (language, word) or append."""
    if not path.exists():
        append_entry(path, entry, columns)
        return

    wb = _load_for_write(path)
    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        append_entry(path, entry, columns)
        return

    ws = wb[SHEET_NAME]
    file_columns = read_header(ws) or columns
    target_row = _find_row(ws, file_columns, entry)
    values = [_render_cell(entry, key) for key in file_columns]
    if target_row is None:
        ws.append(values)
        _style_last_row(ws, file_columns)
    else:
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=target_row, column=col_idx, value=value)
        _style_row(ws, target_row, file_columns)
    _save(wb, path)


def delete_entries(path: Path, language: str, word_keys: set[str]) -> int:
    """Delete rows whose (language, normalized word) is in word_keys.

    Returns the number of rows removed. Preserves header styling and
    sheet structure.
    """
    if not path.exists() or not word_keys:
        return 0
    wb = _load_for_write(path)
    if SHEET_NAME not in wb.sheetnames:
        return 0
    ws = wb[SHEET_NAME]
    columns = read_header(ws)
    if "language" not in columns or "word" not in columns:
        return 0
    lang_idx = columns.index("language") + 1
    word_idx = columns.index("word") + 1

    targets: list[int] = []
    for row in range(2, ws.max_row + 1):
        lang_val = ws.cell(row=row, column=lang_idx).value
        word_val = ws.cell(row=row, column=word_idx).value
        if lang_val != language or not isinstance(word_val, str):
            continue
        key = normalize_word_key(word_val, language)  # type: ignore[arg-type]
        if key in word_keys:
            targets.append(row)
    # Delete bottom-up so indices stay valid.
    for row in reversed(targets):
        ws.delete_rows(row, 1)
    _save(wb, path)
    return len(targets)


def save_with_resolver(
    path: Path,
    entry: VocabularyEntry,
    columns: list[str],
    resolver,
):
    """Single-workbook-load save path with explicit action protocol.

    `resolver(existing_entry_or_None, candidate_entry)` must return a
    2-tuple ``(action, entry_to_write)`` where action is one of:
      - "create"      — no duplicate found; append the candidate
      - "overwrite"   — duplicate found; replace the existing row
      - "append_new"  — duplicate found; append as an extra row
      - "skip"        — duplicate found; keep existing, write nothing

    Returns ``(action, written_entry)``. For "skip", written_entry is the
    existing row read back from the sheet.

    Workbook is opened once, mutated in-place, and saved exactly once.
    """
    if not path.exists():
        ensure_workbook(path, columns)

    wb = _load_for_write(path)
    if SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_NAME)
        _write_header(ws, columns)
    else:
        ws = wb[SHEET_NAME]

    file_columns = read_header(ws) or columns
    if not read_header(ws):
        _write_header(ws, file_columns)

    existing_row = _find_row(ws, file_columns, entry)
    existing_entry: VocabularyEntry | None = None
    if existing_row is not None:
        raw = tuple(
            ws.cell(row=existing_row, column=col_idx).value
            for col_idx in range(1, len(file_columns) + 1)
        )
        existing_entry = _row_to_entry(file_columns, raw)

    action, resolved = resolver(existing_entry, entry)

    if action == "skip":
        # Workbook unchanged — don't bother saving.
        return action, existing_entry or entry

    values = [_render_cell(resolved, key) for key in file_columns]
    if action == "overwrite" and existing_row is not None:
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=existing_row, column=col_idx, value=value)
        _style_row(ws, existing_row, file_columns)
    else:
        # "create" or "append_new" both result in a new row at the end.
        ws.append(values)
        _style_last_row(ws, file_columns)
    _save(wb, path)
    return action, resolved


# ---------- internal helpers (private) ----------------------------


def _load_for_write(path: Path) -> Workbook:
    try:
        return load_workbook(path)
    except OSError as exc:
        raise ExcelLockedError(str(exc)) from exc
    except Exception as exc:
        raise ExcelFormatError(str(exc)) from exc


def _save(wb: Workbook, path: Path) -> None:
    try:
        wb.save(path)
    except PermissionError as exc:
        raise ExcelLockedError(str(exc)) from exc
    except OSError as exc:
        raise StorageError(str(exc)) from exc


def _write_header(ws, columns: list[str]) -> None:
    labels = [COLUMN_LABELS.get(k, k) for k in columns]
    ws.append(labels)
    for idx, key in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(idx)].width = COLUMN_WIDTHS.get(key, 20)
    ws.freeze_panes = "A2"


def _style_last_row(ws, columns: list[str]) -> None:
    _style_row(ws, ws.max_row, columns)


def _style_row(ws, row: int, columns: list[str]) -> None:
    for col_idx, key in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if key == "source_url" and isinstance(cell.value, str) and cell.value.startswith(("http://", "https://")):
            cell.hyperlink = cell.value
            cell.style = "Hyperlink"


def _find_row(ws, columns: list[str], entry: VocabularyEntry) -> int | None:
    if "language" not in columns or "word" not in columns:
        return None
    lang_idx = columns.index("language") + 1
    word_idx = columns.index("word") + 1

    target_key = entry.word_key()
    for row in range(2, ws.max_row + 1):
        lang_val = ws.cell(row=row, column=lang_idx).value
        word_val = ws.cell(row=row, column=word_idx).value
        if lang_val != entry.language or not isinstance(word_val, str):
            continue
        if normalize_word_key(word_val, entry.language) == target_key:
            return row
    return None

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from app.core.models import (
    Example,
    MeaningGroup,
    Sense,
    SubSense,
    VocabularyEntry,
    build_meanings_summary,
    collect_examples_flat,
)
from app.storage.excel_writer import (
    SHEET_NAME,
    append_entry,
    ensure_workbook,
    find_existing,
    update_or_append,
)
from app.storage.settings_store import EXCEL_COLUMN_KEYS_DEFAULT


def _entry_apple() -> VocabularyEntry:
    entry = VocabularyEntry(
        language="en",
        word="apple",
        reading="/ˈæp.əl/",
        part_of_speech=["noun"],
        meaning_groups=[
            MeaningGroup(
                pos="noun",
                senses=[
                    Sense(
                        number=1,
                        gloss="a round fruit",
                        sub_senses=[
                            SubSense(
                                gloss="the fruit",
                                examples=[
                                    Example(
                                        source_text="I ate an apple.",
                                        source_text_plain="I ate an apple.",
                                        translation_ko="나는 사과를 먹었다.",
                                    )
                                ],
                            )
                        ],
                    )
                ],
            )
        ],
        memo="first",
        source_url="https://en.dict.naver.com/#/entry/enko/x",
    )
    entry.examples_flat = collect_examples_flat(entry)
    entry.meanings_summary = build_meanings_summary(entry)
    return entry


def test_ensure_workbook_creates_with_header(tmp_path: Path):
    path = tmp_path / "vocab.xlsx"
    ensure_workbook(path, EXCEL_COLUMN_KEYS_DEFAULT)
    assert path.exists()
    wb = load_workbook(path)
    assert SHEET_NAME in wb.sheetnames
    ws = wb[SHEET_NAME]
    assert ws.cell(row=1, column=1).value == "Language"


def test_append_entry_writes_row(tmp_path: Path):
    path = tmp_path / "vocab.xlsx"
    append_entry(path, _entry_apple(), EXCEL_COLUMN_KEYS_DEFAULT)

    wb = load_workbook(path)
    ws = wb[SHEET_NAME]
    assert ws.max_row == 2
    assert ws.cell(row=2, column=1).value == "en"
    assert ws.cell(row=2, column=2).value == "apple"


def test_find_existing_normalizes_key(tmp_path: Path):
    path = tmp_path / "vocab.xlsx"
    append_entry(path, _entry_apple(), EXCEL_COLUMN_KEYS_DEFAULT)
    found = find_existing(path, "en", "apple")
    assert found is not None
    assert found.word == "apple"


def test_update_or_append_replaces_row(tmp_path: Path):
    path = tmp_path / "vocab.xlsx"
    append_entry(path, _entry_apple(), EXCEL_COLUMN_KEYS_DEFAULT)
    updated = _entry_apple()
    updated.memo = "second"
    update_or_append(path, updated, EXCEL_COLUMN_KEYS_DEFAULT)

    wb = load_workbook(path)
    ws = wb[SHEET_NAME]
    assert ws.max_row == 2  # not appended
    memo_col = EXCEL_COLUMN_KEYS_DEFAULT.index("memo") + 1
    assert ws.cell(row=2, column=memo_col).value == "second"


def test_list_entries_returns_rows(tmp_path: Path):
    path = tmp_path / "vocab.xlsx"
    append_entry(path, _entry_apple(), EXCEL_COLUMN_KEYS_DEFAULT)

    from app.storage.excel_writer import list_entries

    entries = list_entries(path)
    assert len(entries) == 1
    assert entries[0].word == "apple"


def test_delete_entries_removes_matching_rows(tmp_path: Path):
    from app.storage.excel_writer import delete_entries, list_entries

    path = tmp_path / "vocab.xlsx"
    append_entry(path, _entry_apple(), EXCEL_COLUMN_KEYS_DEFAULT)
    append_entry(path, VocabularyEntry(language="en", word="banana"), EXCEL_COLUMN_KEYS_DEFAULT)

    removed = delete_entries(path, "en", {"apple"})
    assert removed == 1
    remaining = list_entries(path)
    assert [e.word for e in remaining] == ["banana"]


def test_delete_entries_noop_when_keys_missing(tmp_path: Path):
    from app.storage.excel_writer import delete_entries

    path = tmp_path / "vocab.xlsx"
    append_entry(path, _entry_apple(), EXCEL_COLUMN_KEYS_DEFAULT)

    assert delete_entries(path, "en", {"banana"}) == 0


def test_update_or_append_appends_when_missing(tmp_path: Path):
    path = tmp_path / "vocab.xlsx"
    append_entry(path, _entry_apple(), EXCEL_COLUMN_KEYS_DEFAULT)
    other = VocabularyEntry(language="en", word="banana")
    update_or_append(path, other, EXCEL_COLUMN_KEYS_DEFAULT)

    wb = load_workbook(path)
    ws = wb[SHEET_NAME]
    assert ws.max_row == 3

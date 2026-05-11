"""Bundles current Excel rows into Anki TSV/APKG output."""
from __future__ import annotations

import logging
from copy import deepcopy
from pathlib import Path

from openpyxl import load_workbook

from app.anki import apkg_exporter, tsv_exporter
from app.core.models import (
    Example,
    MeaningGroup,
    Sense,
    SubSense,
    VocabularyEntry,
    build_meanings_summary,
)
from app.storage import cache_store
from app.storage.excel_serializer import SHEET_NAME, label_to_key
from app.storage.settings_store import Settings

log = logging.getLogger(__name__)


class ExportService:
    def __init__(self, settings: Settings, cache: cache_store.CacheStore) -> None:
        self._settings = settings
        self._cache = cache

    def export_tsv(self, output_path: Path, language: str, deck_name: str | None = None) -> int:
        entries = list(self._collect_entries(language))
        return tsv_exporter.export_tsv(output_path, entries)

    def export_apkg(
        self,
        output_path: Path,
        deck_name: str,
        language: str,
        progress_callback=None,
    ) -> int:
        entries = list(self._collect_entries(language))
        return apkg_exporter.export_apkg(
            output_path, entries, deck_name,
            settings=self._settings,
            progress_callback=progress_callback,
        )

    def _collect_entries(self, language: str) -> list[VocabularyEntry]:
        path = self._excel_path(language)
        if path is None or not path.exists():
            return []

        wb = load_workbook(path, read_only=True)
        if SHEET_NAME not in wb.sheetnames:
            return []
        ws = wb[SHEET_NAME]
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            return []
        keys = [label_to_key(label) for label in header]

        entries: list[VocabularyEntry] = []
        for raw in rows:
            if raw is None:
                continue
            data = dict(zip(keys, [v if v is not None else "" for v in raw]))
            if not data.get("word"):
                continue
            entries.append(self._row_to_entry(data))
        return entries

    def _excel_path(self, language: str) -> Path | None:
        path_str = self._settings.excel_path_for(language)
        return Path(path_str) if path_str else None

    def _row_to_entry(self, data: dict) -> VocabularyEntry:
        word = str(data.get("word", "")).strip()
        language = str(data.get("language", "en")).strip() or "en"
        cached = self._cache.get(word, language) if self._cache else None  # type: ignore[arg-type]
        return entry_from_export_row(data, cached)


def _split_csv(value) -> list[str]:
    if not value:
        return []
    return [p.strip() for p in str(value).split(",") if p.strip()]


def _entry_from_flat_row(data: dict) -> VocabularyEntry:
    word = str(data.get("word", "")).strip()
    language = str(data.get("language", "en")).strip() or "en"
    sources = [s for s in str(data.get("examples", "") or "").split("\n") if s]
    translations = [s for s in str(data.get("example_translations", "") or "").split("\n") if s]
    examples = []
    for idx, src in enumerate(sources):
        examples.append(
            Example(
                source_text=src,
                source_text_plain=src,
                translation_ko=translations[idx] if idx < len(translations) else None,
                order=idx,
            )
        )

    pos_list = _split_csv(data.get("part_of_speech", ""))
    summary = str(data.get("meanings_summary", "") or "")
    meaning_groups: list[MeaningGroup] = []
    if pos_list and summary:
        meaning_groups = [
            MeaningGroup(
                pos=pos_list[0],
                senses=[Sense(number=1, gloss=summary, sub_senses=[SubSense(examples=examples)])],
            )
        ]

    entry = VocabularyEntry(
        language=language,  # type: ignore[arg-type]
        word=word,
        reading=str(data.get("reading", "") or "") or None,
        part_of_speech=pos_list,
        meaning_groups=meaning_groups,
        meanings_summary=summary,
        examples_flat=examples,
        synonyms=_split_csv(data.get("synonyms", "")),
        antonyms=_split_csv(data.get("antonyms", "")),
        tags=_split_csv(data.get("tags", "")),
        memo=str(data.get("memo", "") or ""),
        source_url=str(data.get("source_url", "") or "") or None,
        source_provider="unknown",
        created_at=str(data.get("created_at", "") or ""),
        updated_at=str(data.get("updated_at", "") or ""),
    )
    if not entry.meanings_summary:
        entry.meanings_summary = build_meanings_summary(entry)
    return entry


def entry_from_export_row(
    data: dict,
    cached: VocabularyEntry | None = None,
) -> VocabularyEntry:
    """Build an export entry with Excel as the editable source of truth.

    The cache is allowed to restore rich nested meaning/example structure,
    but it must not overwrite cells the user can edit in Excel. To avoid
    exporting stale definitions after a spreadsheet edit, cached meaning
    groups are reused only when the visible summary still matches.
    """
    base = _entry_from_flat_row(data)
    if cached is None:
        return base

    cached_summary = cached.meanings_summary or build_meanings_summary(cached)
    base_summary = str(data.get("meanings_summary", "") or base.meanings_summary or "")
    if cached.meaning_groups and (not base_summary or base_summary == cached_summary):
        base.meaning_groups = deepcopy(cached.meaning_groups)
        if base.examples_flat:
            _replace_nested_examples(base.meaning_groups, base.examples_flat)
        if not base.meanings_summary:
            base.meanings_summary = cached_summary

    if not base.examples_flat and cached.examples_flat:
        base.examples_flat = deepcopy(cached.examples_flat)
        for group in base.meaning_groups:
            for sense in group.senses:
                for sub in sense.sub_senses:
                    if not sub.examples:
                        sub.examples = deepcopy(cached.examples_flat)

    base.source_provider = "excel"
    return base


def _replace_nested_examples(
    groups: list[MeaningGroup],
    examples: list[Example],
) -> None:
    attached = False
    for group in groups:
        for sense in group.senses:
            for sub in sense.sub_senses:
                if not attached:
                    sub.examples = deepcopy(examples)
                    attached = True
                else:
                    sub.examples = []
    if attached or not groups or not groups[0].senses:
        return
    groups[0].senses[0].sub_senses.append(SubSense(examples=deepcopy(examples)))
